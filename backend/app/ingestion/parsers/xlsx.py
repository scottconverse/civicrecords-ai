import logging
import tempfile
import zipfile
from pathlib import Path

from openpyxl import load_workbook

from app.ingestion.parsers.base import BaseParser, ParsedPage, ParseResult

logger = logging.getLogger(__name__)

# ZIP entries that indicate embedded VBA macros in XLSX files.
_XLSX_VBA_ENTRIES = {"xl/vbaproject.bin"}


class XlsxParser(BaseParser):
    supported_extensions = [".xlsx", ".xlsm", ".xls"]

    @staticmethod
    def _strip_macros(file_path: Path) -> tuple[Path | None, list[str]]:
        """Check for VBA macro entries and return a sanitized copy if found.

        Returns (sanitized_path, stripped_entry_names).
        sanitized_path is None if no macros were found.
        """
        try:
            with zipfile.ZipFile(file_path, "r") as zin:
                found = [n for n in zin.namelist() if n.lower() in _XLSX_VBA_ENTRIES]
                if not found:
                    return None, []
                tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
                tmp_path = Path(tmp.name)
                tmp.close()
                with zipfile.ZipFile(tmp_path, "w", zipfile.ZIP_DEFLATED) as zout:
                    for item in zin.namelist():
                        if item.lower() not in _XLSX_VBA_ENTRIES:
                            zout.writestr(item, zin.read(item))
                logger.warning(
                    "Stripped VBA macros from %s: %s", file_path.name, found
                )
                return tmp_path, found
        except zipfile.BadZipFile:
            logger.warning("Not a valid ZIP archive: %s", file_path.name)
            return None, []

    def parse(self, file_path: Path) -> ParseResult:
        sanitized_path, stripped_entries = self._strip_macros(file_path)
        parse_path = sanitized_path or file_path
        try:
            wb = load_workbook(str(parse_path), read_only=True, data_only=True)
            pages = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = []
                for row in ws.iter_rows(values_only=True):
                    cells = [str(c) if c is not None else "" for c in row]
                    if any(c.strip() for c in cells):
                        rows.append(" | ".join(cells))
                if rows:
                    text = f"Sheet: {sheet_name}\n" + "\n".join(rows)
                    pages.append(ParsedPage(
                        text=text, page_number=None, metadata={"sheet": sheet_name}
                    ))
            sheet_count = len(wb.sheetnames)
            wb.close()
            metadata: dict = {"sheet_count": sheet_count}
            if stripped_entries:
                metadata["macros_stripped"] = True
                metadata["stripped_entries"] = stripped_entries
            return ParseResult(pages=pages, metadata=metadata, file_type="xlsx")
        finally:
            if sanitized_path and sanitized_path.exists():
                sanitized_path.unlink()
